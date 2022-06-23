import openpyxl as xl
path = 'C:\\Programming\\python-tutorial\\01-hello-world\\Project 1\\transactions.xlsx'
wb = xl.load_workbook(path)
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1,1)
# print(cell.value)
# print(sheet.max_row)

# for row in range (2, sheet.max_row + 1):
    # cell = sheet.cell(row,3)
    # print(cell.value)

for row in range (2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

wb.save('Project 1/transactions2.xlsx')